import os
import sys

import openpyxl
from PyQt6.QtCore import QTimer
from PySide6.QtGui import QIcon, QFont, QGuiApplication
from PySide6.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QWidget, QComboBox, QLabel, QLineEdit, \
    QPushButton, QMessageBox, QHBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView
import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill


class fAlimentacao(QMainWindow):
    last_creation_date = None
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Cadastro de Funcionários")
        self.setGeometry(600, 200, 700, 700)
        self.setStyleSheet("background-color:blue; color:white; border-radius: 10px;")
        self.setFont(QFont("Times", 20))
        self.setWindowIcon(QIcon("C:/Users/fredi/PycharmProjects/pythonProject/logo-wasion.png"))

        # Configuração da janela principal
        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)

        # Layout vertical para o widget central
        layout = QVBoxLayout(central_widget)

    def showEvent(self, event):
        # Quando a janela é mostrada, maximiza ela automaticamente
        self.showMaximized()
        event.accept()

        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)

        # Conecta ao banco de dados modelo
        self.conn_modelo = sqlite3.connect("modelo.db")
        self.cursor_modelo = self.conn_modelo.cursor()

        # Conecta ao banco de dados funcionários
        self.conn_funcionario = sqlite3.connect("funcionario.db")
        self.cursor_funcionario = self.conn_funcionario.cursor()

        # Cria um layout principal
        layout = QVBoxLayout()

        # Cria uma lista suspensa (QComboBox Modelo)
        self.combobox_modelos = QComboBox()
        self.combobox_modelos.setFont(QFont("Times", 20))
        self.combobox_modelos.setStyleSheet("background-color: white; color: black; border-radius:10px;")
        self.combobox_modelos.addItem("")  # Linha em branco
        self.carregar_modelos()
        self.combobox_modelos.currentIndexChanged.connect(self.abrir_nova_janela)

        # Cria uma lista suspensa (QComboBox Funcionário)
        self.combobox_funcionario = QComboBox()
        self.combobox_funcionario.setFont(QFont("Times", 20))
        self.combobox_funcionario.setStyleSheet("background-color: white; color: black; border-radius: 10px;")
        self.combobox_funcionario.addItem("")
        self.carregar_funcionarios()
        self.combobox_funcionario.currentIndexChanged.connect(self.abrir_nova_janela)

        self.combobox_turno = QComboBox()
        self.combobox_turno.setFont(QFont("Times", 20))
        self.combobox_turno.setStyleSheet("background-color: white; color: black; border-radius: 10px;")
        self.combobox_turno.addItem("")  # Linha em branco
        # Adicione os turnos disponíveis (você pode alterar os itens conforme necessário)
        self.combobox_turno.addItem("1")
        self.combobox_turno.addItem("2")
        self.combobox_turno.addItem("3")
        self.combobox_turno.addItem("Comercial")
        # Conecte um slot (método) para lidar com a mudança do valor selecionado
        self.combobox_turno.currentIndexChanged.connect(self.abrir_nova_janela)

        
        # Adiciona os QComboBoxes ao layout
        layout.addWidget(QLabel("Modelo:"))
        layout.addWidget(self.combobox_modelos)
        layout.addWidget(QLabel("Funcionário:"))
        layout.addWidget(self.combobox_funcionario)
        layout.addWidget(QLabel("Turno:"))  # Adicione esta linha
        layout.addWidget(self.combobox_turno)

        # Cria um widget central e define o layout principal
        central_widget = QWidget()
        central_widget.setLayout(layout)
        self.setCentralWidget(central_widget)

    def carregar_modelos(self):
        self.cursor_modelo.execute("SELECT name FROM 'sqlite_master' WHERE type='table' AND name != 'sqlite_sequence'")
        tabelas = self.cursor_modelo.fetchall()
        for tabela in tabelas:
            self.combobox_modelos.addItem(tabela[0])

    def carregar_funcionarios(self):
        self.cursor_funcionario.execute("SELECT nome FROM funcionarios")
        funcionarios = self.cursor_funcionario.fetchall()
        for funcionario in funcionarios:
            self.combobox_funcionario.addItem(funcionario[0])

    def abrir_nova_janela(self, index):
        modelo = self.combobox_modelos.currentText()
        funcionarios = self.combobox_funcionario.currentText()
        turno = self.combobox_turno.currentText()
        if modelo and funcionarios and turno:
            self.hide()
            self.new_janela = janelaAlimentacao(modelo, funcionarios, turno)
            self.new_janela.show()


class janelaAlimentacao(QWidget):
    def __init__(self, modelo, funcionarios=None, turno=None):
        super().__init__()

        self.cadastrar = None
        self.setWindowTitle("Registro de Alimentação")
        self.setGeometry(600, 200, 700, 700)
        self.setWindowIcon(QIcon('c:/Users/fredi/PycharmProjects/pythonProject/logo-wasion.png'))
        self.setStyleSheet("background-color: blue; color: white;")
        # Maximizar a janela
        self.showMaximized()

        self.conn = sqlite3.connect("modelo.db")
        self.cursor = self.conn.cursor()

        self.modelo = modelo
        self.funcionario = funcionarios
        self.turno = turno

        layout = QVBoxLayout()

        # Modelo
        modelo_label = QLabel(f"Modelo: {self.modelo}")
        modelo_label.setFont(QFont("Times", 20))
        modelo_label.setStyleSheet("color: white")
        layout.addWidget(modelo_label)

        funcionario_label = QLabel(f"Funcionário: {self.funcionario}")
        funcionario_label.setFont(QFont("Times", 20))
        funcionario_label.setStyleSheet("color: white")
        layout.addWidget(funcionario_label)

        turno_label = QLabel(f"Turno: {self.turno}")
        turno_label.setFont(QFont("Times", 20))
        turno_label.setStyleSheet("color: white")
        layout.addWidget(turno_label)

        # Máquina
        self.maquina_edit = QLineEdit()
        self.maquina_edit.setStyleSheet("background-color: white; color: black; border-radius: 10px;")
        self.maquina_edit.setFont(QFont("Times", 20))
        self.maquina_edit.textChanged.connect(self.buscar_codigo)
        layout.addWidget(QLabel("Máquina:"))
        layout.addWidget(self.maquina_edit)

        # Portal
        self.portal_edit = QLineEdit()
        self.portal_edit.setFont(QFont("Times", 20))
        self.portal_edit.setStyleSheet("background-color: white; color: black; border-radius: 10px;")
        self.portal_edit.textChanged.connect(self.buscar_codigo)
        layout.addWidget(QLabel("Portal:"))
        layout.addWidget(self.portal_edit)

        # Posição
        self.posicao_edit = QLineEdit()
        self.posicao_edit.setFont(QFont("Times", 20))
        self.posicao_edit.setStyleSheet("background-color: white; color: black; border-radius: 10px;")
        self.posicao_edit.textChanged.connect(self.buscar_codigo)
        layout.addWidget(QLabel("Posição:"))
        layout.addWidget(self.posicao_edit)

        # Divisão
        self.divisao_edit = QLineEdit()
        self.divisao_edit.setFont(QFont("Times", 20))
        self.divisao_edit.setStyleSheet("background-color: white; color: black; border-radius: 10px;")
        self.divisao_edit.textChanged.connect(self.buscar_codigo)
        layout.addWidget(QLabel("Divisão:"))
        layout.addWidget(self.divisao_edit)

        # Código de Entrada
        self.codigo2_edit = QLineEdit()
        self.codigo2_edit.setFont(QFont("Times", 20))
        self.codigo2_edit.setStyleSheet("background-color: white; color: black; border-radius: 10px;")
        self.codigo2_edit.returnPressed.connect(self.verificar_codigo)
        layout.addWidget(QLabel("Código de Entrada:"))
        layout.addWidget(self.codigo2_edit)

        # Código
        self.codigo1_edit = QLineEdit()
        self.codigo1_edit.setFont(QFont("Times", 20))
        self.codigo1_edit.setStyleSheet("background-color: blue; color: white; border-radius: 10px;")
        self.codigo1_edit.setReadOnly(True)  # Impede que o campo seja editado diretamente pelo usuário
        layout.addWidget(QLabel("Código Correto:"))
        layout.addWidget(self.codigo1_edit)

        # Botões Cadastrar e limpar
        button_layout = QHBoxLayout()
        cadastrar_button = QPushButton("Cadastrar")
        cadastrar_button.setFont(QFont("Times", 20))
        cadastrar_button.setStyleSheet("background-color: green; color: white; border-radius: 10px;")
        cadastrar_button.clicked.connect(self.verificar_codigo)
        limpar_button = QPushButton("Limpar")
        limpar_button.setFont(QFont("Times", 20))
        limpar_button.setStyleSheet("background-color: green; color: white; border-radius: 10px;")
        limpar_button.clicked.connect(self.limpar)
        button_layout.addWidget(cadastrar_button)
        button_layout.addWidget(limpar_button)
        button_layout.addStretch(1)
        layout.addLayout(button_layout)

        # Botão Sair
        sair_button = QPushButton("Sair")
        sair_button.setFont(QFont("Times", 20))
        sair_button.setStyleSheet("background-color: red; color: white; border-radius: 10px;")
        sair_button.clicked.connect(self.voltar_para_fAlimentacao)
        button_layout.addWidget(sair_button)

        # Pré-visualização da planilha
        self.table_preview = QTableWidget()
        self.table_preview.setStyleSheet("background-color: white; color: black")
        self.table_preview.setRowCount(0)
        self.table_preview.setColumnCount(13)
        self.table_preview.setHorizontalHeaderLabels(
            ["Modelo", "Funcionário", "Máquina", "Portal", "Posição", "Divisão", "Código1", "Código2", "Status",
             "Operador", "Qualidade", "Data", "Hora"])
        header = self.table_preview.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        header.setStretchLastSection(True)
        layout.addWidget(self.table_preview)

        self.setLayout(layout)

    def _move_cursor_if_length(self, source, destination, length):
        if len(source.text()) >= length:
            destination.setFocus()

    def buscar_codigo(self):
        maquina = self.maquina_edit.text().strip()
        portal = self.portal_edit.text().strip()
        posicao = self.posicao_edit.text().strip()
        divisao = self.divisao_edit.text().strip()

        if maquina and portal and posicao and divisao:
            querry = "SELECT Codigo1 FROM '{}' WHERE Maquina='{}' AND Portal='{}' AND Posicao='{}' AND Divisao='{}'".format(
                self.modelo, maquina, portal, posicao, divisao)
            self.cursor.execute(querry)
            resultado = self.cursor.fetchone()
            if resultado:
                self.codigo1_edit.setText(resultado[0])
            else:
                self.codigo1_edit.clear()

        # Chamar função _move_cursor_if_length para cada campo
        self._move_cursor_if_length(self.maquina_edit, self.portal_edit, 3)
        self._move_cursor_if_length(self.portal_edit, self.posicao_edit, 1)
        self._move_cursor_if_length(self.posicao_edit, self.divisao_edit, 2)
        self._move_cursor_if_length(self.divisao_edit, self.codigo2_edit, 4)

    def verificar_campos_obrigatorios(self):
        if self.codigo2_edit.text() != "":
            if (
                    self.maquina_edit.text().strip() == "" or
                    self.portal_edit.text().strip() == "" or
                    self.posicao_edit.text().strip() == "" or
                    self.divisao_edit.text().strip() == "" or
                    self.codigo2_edit.text().strip() == ""
            ):
                msg_box = QMessageBox()
                msg_box.setStyleSheet("QMessageBox { background-color: red; color: white; }")
                msg_box.setWindowTitle("Erro")
                msg_box.setText("Preencha todos os campos obrigatórios.")
                msg_box.setFont(QFont("Times", 25))
                msg_box.setWindowIcon(QIcon("C:/Users/fredi/PycharmProjects/pythonProject/thumbs-down.svg"))
                msg_box.exec()
                return False
            return True
        return False

    def verificar_codigo(self):
        if not self.verificar_campos_obrigatorios():
            return

        codigo1 = self.codigo1_edit.text().strip()
        codigo2 = self.codigo2_edit.text().strip()
        cadastrado_com_sucesso = codigo1 == codigo2
        self.exportar_excel(cadastrado_com_sucesso)

        if cadastrado_com_sucesso:
            self.cadastrar = True
            # Exibir caixa de mensagem de cadastro realizado
            msg_box = QMessageBox()
            msg_box.setStyleSheet("QMessageBox { background-color: green; color: white; }")
            msg_box.setWindowTitle("Sucesso")
            msg_box.setText("Código cadastrado com sucesso.")
            msg_box.setFont(QFont("Times", 20))
            msg_box.setWindowIcon(QIcon("thumbs-up.svg"))

            # Remover botão "OK" e tornar a mensagem permanente
            msg_box.setStandardButtons(QMessageBox.NoButton)
            self.success_msg_box = msg_box  # Armazena a referência à caixa de mensagem

            msg_box.show()  # Exibe a caixa de mensagem

            # Agendar a função para fechar a caixa de mensagem após 3 segundos (3000 milissegundos)
            QTimer.singleShot(3000, self.close_success_message)
        else:
            self.cadastrar = False
            msg_box = QMessageBox()
            msg_box.setStyleSheet("QMessageBox { background-color: red; color: white; }")
            msg_box.setWindowTitle("Erro")
            msg_box.setText("Os códigos não conferem")
            msg_box.setFont(QFont("Times", 25))
            msg_box.setWindowIcon(QIcon("C:/Users/fredi/PycharmProjects/pythonProject/thumbs-down.svg"))
            msg_box.setStandardButtons(QMessageBox.Ok)
            msg_box.exec()

            self.codigo2_edit.clear()

    def close_success_message(self):
        if self.success_msg_box is not None:
            self.success_msg_box.close()
            self.success_msg_box = None

            self.maquina_edit.clear()
            self.divisao_edit.clear()
            self.portal_edit.clear()
            self.codigo1_edit.clear()
            self.codigo2_edit.clear()
            self.posicao_edit.clear()
            # Move o foco do cursor para o campo de Máquina
            self.maquina_edit.setFocus()


    def close_success_message(self):
        if self.success_msg_box is not None:
            self.success_msg_box.close()
            self.success_msg_box = None

            self.maquina_edit.clear()
            self.divisao_edit.clear()
            self.portal_edit.clear()
            self.codigo1_edit.clear()
            self.codigo2_edit.clear()
            self.posicao_edit.clear()
            # Move o foco do cursor para o campo de Máquina
            self.maquina_edit.setFocus()
        else:
            self.cadastrar = False
            msg_box = QMessageBox()
            msg_box.setStyleSheet("QMessageBox { background-color: yellow; color: white; }")
            msg_box.setWindowTitle("Erro")
            msg_box.setText("Os códigos não conferem")
            msg_box.setFont(QFont("Times", 25))
            msg_box.setWindowIcon(QIcon("C:/Users/fredi/PycharmProjects/pythonProject/thumbs-down.svg"))
            msg_box.exec()

            self.codigo2_edit.clear()

    def exportar_excel(self, cadastrado_com_sucesso):
        data = {
            "Modelo": [self.modelo],
            "Funcionário": [self.funcionario],
            "Turno": [self.turno],
            "Máquina": [self.maquina_edit.text().strip()],
            "Portal": [self.portal_edit.text().strip()],
            "Posição": [self.posicao_edit.text().strip()],
            "Divisão": [self.divisao_edit.text().strip()],
            "Código1": [self.codigo1_edit.text().strip()],
            "Código2": [self.codigo2_edit.text().strip()],
            "Status": ["APROVADO" if cadastrado_com_sucesso else "REPROVADO"],
            "Qualidade": ["PENDENTE"],
            "Data": [datetime.now().strftime("%d/%m/%Y")],  # Salvar a data em formato "Dia/Mes/Ano"
            "Hora": [datetime.now().strftime("%H:%M:%S")],  # Salvar a hora em formato "HH:MM:SS"
        }

        df = pd.DataFrame(data)

        # Obtém a data atual no formato 'dd-mm-aaaa'
        dia_atual = datetime.now().strftime('%d-%m-%Y')

        # Define o nome do arquivo usando o nome do modelo e a data atual
        nome_arquivo = f"{self.modelo}_{dia_atual}.xlsx"

        # Verifica se a data atual é diferente da última data de criação
        if fAlimentacao.last_creation_date != dia_atual:
            fAlimentacao.last_creation_date = dia_atual

            # Cria ou obtém a planilha ativa
            wb = openpyxl.Workbook()
            ws = wb.active

            # Adiciona os campos como cabeçalho apenas se for uma nova planilha
            for col, field in enumerate(df.columns, start=1):
                ws.cell(row=1, column=col, value=field)

            # Define as larguras das colunas e formatação
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 15
            ws.column_dimensions["F"].width = 20
            ws.column_dimensions["G"].width = 20
            ws.column_dimensions["H"].width = 20
            ws.column_dimensions["I"].width = 20
            ws.column_dimensions["J"].width = 20
            ws.column_dimensions["K"].width = 20
            ws.column_dimensions["L"].width = 15  # Adicionando largura para a coluna Data
            ws.column_dimensions["M"].width = 15  # Adicionando largura para a coluna Hora

            for cell in ws["A"]:
                cell.alignment = Alignment(horizontal="center")

            data_hora_col = ws["H"]
            for cell in data_hora_col[1:]:
                cell.number_format = "dd/mm/yyyy HH:mm:ss"

        else:
            # Se a data for igual à última data de criação, carrega a planilha existente
            if os.path.exists(nome_arquivo):
                wb = openpyxl.load_workbook(nome_arquivo)
                ws = wb.active
            else:
                # Caso contrário, cria uma nova planilha
                wb = openpyxl.Workbook()
                ws = wb.active

        # Escreve os dados na planilha
        for row_index, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start=ws.max_row + 1):
            for col_index, cell_data in enumerate(row_data, start=1):
                ws.cell(row=row_index, column=col_index, value=cell_data)

        # Aplica cores às linhas da planilha de acordo com o status (Aprovado ou Reprovado)
        for row in ws.iter_rows(min_row=2, min_col=9, max_col=10):
            for cell in row:
                if cell.value == "APROVADO":
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Salva o arquivo XLSX
        wb.save(nome_arquivo)

        # Pré-visualização na interface
        self.preview_planilha(nome_arquivo)

    def preview_planilha(self, nome_arquivo):
        # Carregar os dados da planilha para a tabela de pré-visualização
        wb = openpyxl.load_workbook(nome_arquivo)
        ws = wb.active
        data = list(ws.iter_rows(values_only=True))
        headers = data[0]
        data = data[1:]
        self.table_preview.setRowCount(len(data))
        self.table_preview.setColumnCount(len(headers))
        self.table_preview.setHorizontalHeaderLabels(headers)
        for row_index, row_data in enumerate(data):
            for col_index, cell_data in enumerate(row_data):
                item = QTableWidgetItem(str(cell_data))
                self.table_preview.setItem(row_index, col_index, item)

    def voltar_para_fAlimentacao(self):
        # Fechar a janela atual (janelaAlimentacao)
        self.close()
        # Criar uma instância da classe fAlimentacao
        self.f_alimentacao = fAlimentacao()
        # Exibir novamente a janela fAlimentacao
        self.f_alimentacao.show()

    def limpar(self):
        self.maquina_edit.clear()
        self.divisao_edit.clear()
        self.portal_edit.clear()
        self.codigo1_edit.clear()
        self.codigo2_edit.clear()
        self.posicao_edit.clear()
        self.maquina_edit.setFocus()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    tela_admin = fAlimentacao()
    tela_admin.show()
    sys.exit(app.exec())